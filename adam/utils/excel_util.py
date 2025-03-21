# -*- coding:utf-8 -*-
import os
import csv
import json
import logging

import openpyxl
from .str_util import DECODE_CODING_LIST

LOGGER = logging.getLogger(__name__)


def csv_reader(file_path):
    """csv格式的文件读取
    :param file_path: 文件路径
    :return: 读取结果(按行读取的列表)
    """
    result = []
    null_set = set([''])
    for encode in DECODE_CODING_LIST:
        try:
            with open(file_path, encoding=encode) as f:
                reader = csv.reader(f)
                result = [row for row in reader if set(row) != null_set]
                break
        except UnicodeDecodeError as e:
            pass
    return result


def excel_reader(file_path):
    """
    读取 Excel 内容
    :param {string} file_path: 需读取的 excel 文件路径
    :return {dict}: Excel 内容的dict, {标签页名称:按行列组成的二维数组table}
    """
    # office 2007 文件读取(data_only:是否取公式计算后的值,默认取那条公式)
    workbook = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
    data = {}  # 数据容器,内容为 {标签页名称:内容}
    # active_sheet = workbook.active # 被选中的标签页
    # 遍历各标签页
    for sheet in workbook:
        max_row = sheet.max_row  # 最大行数
        # max_column = sheet.max_column  # 最大列数
        keep_data = []  # 解析类型后的新结果
        # 遍历工作表中的所有行
        for row_num, row_values in enumerate(sheet.iter_rows(values_only=True)):
            # 完全为空的一行数据，不添加进来
            if row_values:
                keep_data.append(row_values)
            if row_num % 100 == 0:  # 每读取100行打一次日志，减少日志量
                LOGGER.info(f'读取数据 {row_num}/{max_row} 完成 {row_num/max_row*100:.2f}%')
        # 没有内容的标签页，不处理
        if not keep_data:
            continue
        data[sheet.title] = keep_data
    return data

